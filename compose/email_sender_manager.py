#!/usr/bin/env python3
"""
콜드메일 발송 관리 시스템
- 기존 이커머스 데이터베이스와 연동
- 발송대기 리스트 엑셀 관리
- 콜드메일 생성 시 자동 업데이트
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from loguru import logger
from core.config import load_config_data
import json


class EmailSenderManager:
    """콜드메일 발송 관리 시스템"""

    def __init__(self):
        # 경로 설정
        config_data = load_config_data()
        file_org = config_data.get('file_organizer', {})
        project_root = Path(__file__).resolve().parents[1]
        default_work = project_root / 'data' / 'work'
        default_db = project_root / 'data' / 'storage' / 'ecommerce_database.xlsx'
        default_coldmail = project_root / 'data' / 'coldmail'

        self.work_folder = Path(file_org.get('work_folder', default_work)).expanduser()
        self.database_file = Path(file_org.get('database_file', default_db)).expanduser()
        self.coldmail_folder = Path(file_org.get('coldmail_folder', default_coldmail)).expanduser()
        self.pending_folder = self.coldmail_folder / '발송대기'
        self.sent_folder = self.coldmail_folder / '발송완료'
        self.pending_excel = self.pending_folder / '콜드메일_발송대기_리스트.xlsx'

        # 폴더 생성
        self.pending_folder.mkdir(parents=True, exist_ok=True)
        self.sent_folder.mkdir(parents=True, exist_ok=True)

        # 발송대기 엑셀 초기화
        self._init_pending_excel()

    def _init_pending_excel(self):
        """발송대기 엑셀 파일 초기화 (가독성 좋은 포맷팅 적용)"""
        if not self.pending_excel.exists():
            # 엑셀 양식 생성
            columns = [
                '등록일시',
                '스토어명',
                '스토어URL',
                '담당자정보',
                '상품명',
                '콜드메일_제목',
                '콜드메일_내용',
                '생성일시',
                '발송예정일',
                '발송상태',
                '발송일시',
                '응답여부',
                '응답내용',
                '메모'
            ]

            df = pd.DataFrame(columns=columns)

            # openpyxl을 사용한 고급 포맷팅 적용
            self._create_formatted_excel(df, columns)
            logger.info(f"발송대기 엑셀 파일 생성 (포맷팅 적용): {self.pending_excel}")

    def _create_formatted_excel(self, df: pd.DataFrame, columns: List[str]):
        """가독성 좋은 엑셀 파일 생성"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.worksheet.table import Table, TableStyleInfo

            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "콜드메일_발송대기"

            # 헤더 스타일 정의
            header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 데이터 스타일 정의
            data_font = Font(name='맑은 고딕', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 테두리 스타일
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 헤더 추가 및 스타일 적용
            for col_idx, column_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 컬럼 너비 설정 (가독성 향상)
            column_widths = {
                'A': 16,  # 등록일시
                'B': 20,  # 스토어명
                'C': 25,  # 스토어URL
                'D': 18,  # 담당자정보
                'E': 25,  # 상품명
                'F': 25,  # 콜드메일_제목
                'G': 40,  # 콜드메일_내용
                'H': 16,  # 생성일시
                'I': 12,  # 발송예정일
                'J': 10,  # 발송상태
                'K': 16,  # 발송일시
                'L': 8,   # 응답여부
                'M': 30,  # 응답내용
                'N': 20   # 메모
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 행 높이 설정
            ws.row_dimensions[1].height = 25  # 헤더 행 높이

            # 기본 데이터 행 스타일 설정 (10행까지 미리 설정)
            for row in range(2, 12):
                ws.row_dimensions[row].height = 60  # 데이터 행 높이 (콘텐츠용)
                for col in range(1, len(columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # 상태별 조건부 서식을 위한 셀 스타일 정의
            status_colors = {
                '대기': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
                '발송완료': PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid'),
                '실패': PatternFill(start_color='F8CECC', end_color='F8CECC', fill_type='solid')
            }

            # 시트 보호 및 필터 설정
            ws.auto_filter.ref = f"A1:{chr(64 + len(columns))}1"

            # 파일 저장
            wb.save(self.pending_excel)

        except ImportError:
            # openpyxl이 없는 경우 기본 엑셀로 저장
            logger.warning("openpyxl이 설치되지 않아 기본 형식으로 저장합니다")
            df.to_excel(self.pending_excel, index=False)
        except Exception as e:
            logger.error(f"포맷팅된 엑셀 생성 실패, 기본 형식으로 저장: {e}")
            df.to_excel(self.pending_excel, index=False)

    def _save_with_formatting(self, df: pd.DataFrame):
        """포맷팅을 유지하면서 데이터 저장"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # 기존 워크북 로드
            wb = load_workbook(self.pending_excel)
            ws = wb.active

            # 기존 데이터 삭제 (헤더 제외)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None

            # 데이터 스타일 정의
            data_font = Font(name='맑은 고딕', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 상태별 색상 정의
            status_colors = {
                '대기': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
                '발송완료': PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid'),
                '실패': PatternFill(start_color='F8CECC', end_color='F8CECC', fill_type='solid')
            }

            # 데이터 입력 및 스타일 적용
            for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
                ws.row_dimensions[row_idx].height = 60  # 행 높이

                for col_idx, (col_name, value) in enumerate(row_data.items(), start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # 발송상태에 따른 색상 적용
                    if col_name == '발송상태' and str(value) in status_colors:
                        # 해당 행 전체에 색상 적용
                        for c in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=c).fill = status_colors[str(value)]

            # 자동 필터 유지
            ws.auto_filter.ref = f"A1:{chr(64 + len(df.columns))}{len(df) + 1}"

            # 저장
            wb.save(self.pending_excel)

        except Exception as e:
            logger.error(f"포맷팅 저장 실패, 기본 저장: {e}")
            # openpyxl 없을 때는 pandas의 ExcelWriter 사용
            with pd.ExcelWriter(self.pending_excel, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='발송대기')

                # 워크시트 접근
                workbook = writer.book
                worksheet = writer.sheets['발송대기']

                # 기본 컬럼 너비 설정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                    worksheet.column_dimensions[column_letter].width = adjusted_width

    def load_ecommerce_database(self) -> pd.DataFrame:
        """기존 이커머스 데이터베이스 로드"""
        try:
            if self.database_file.exists():
                df = pd.read_excel(self.database_file)
                logger.info(f"이커머스 데이터베이스 로드 완료: {len(df)}개 항목")
                return df
            else:
                logger.warning(f"이커머스 데이터베이스 파일이 없습니다: {self.database_file}")
                return pd.DataFrame()
        except Exception as e:
            logger.error(f"이커머스 데이터베이스 로드 실패: {e}")
            return pd.DataFrame()

    def load_pending_list(self) -> pd.DataFrame:
        """발송대기 리스트 로드"""
        try:
            if self.pending_excel.exists():
                df = pd.read_excel(self.pending_excel)
                logger.info(f"발송대기 리스트 로드: {len(df)}개 항목")
                return df
            else:
                return pd.DataFrame()
        except Exception as e:
            logger.error(f"발송대기 리스트 로드 실패: {e}")
            return pd.DataFrame()

    def add_to_pending_list(self,
                          store_name: str,
                          store_url: str = "",
                          contact_info: str = "",
                          product_name: str = "",
                          email_subject: str = "",
                          email_content: str = "",
                          scheduled_date: Optional[datetime] = None) -> bool:
        """발송대기 리스트에 추가"""
        try:
            # 발송대기 리스트 로드
            df = self.load_pending_list()

            # 발송예정일 설정 (기본값: 내일)
            if scheduled_date is None:
                scheduled_date = datetime.now() + timedelta(days=1)

            # 새 항목 생성
            new_row = {
                '등록일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '스토어명': store_name,
                '스토어URL': store_url,
                '담당자정보': contact_info,
                '상품명': product_name,
                '콜드메일_제목': email_subject,
                '콜드메일_내용': email_content,
                '생성일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '발송예정일': scheduled_date.strftime('%Y-%m-%d'),
                '발송상태': '대기',
                '발송일시': '',
                '응답여부': 'N',
                '응답내용': '',
                '메모': ''
            }

            # 데이터프레임에 추가
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # 포맷팅을 유지하면서 엑셀 저장
            self._save_with_formatting(df)
            logger.info(f"발송대기 리스트에 추가: {store_name}")

            return True

        except Exception as e:
            logger.error(f"발송대기 리스트 추가 실패: {e}")
            return False

    def add_coldmail_result(self, coldmail_result: Dict[str, Any]) -> bool:
        """콜드메일 생성 결과를 발송대기 리스트에 자동 추가"""
        try:
            # 콜드메일 결과에서 정보 추출
            store_name = self._extract_store_name(coldmail_result)
            product_name = self._extract_product_name(coldmail_result)
            email_subject = coldmail_result.get('subject', '')
            email_content = coldmail_result.get('body', '')

            # 기존 이커머스 DB에서 스토어 정보 찾기
            ecommerce_df = self.load_ecommerce_database()
            store_info = self._find_store_info(ecommerce_df, store_name, product_name)

            # 발송대기 리스트에 추가
            return self.add_to_pending_list(
                store_name=store_name,
                store_url=store_info.get('store_url', ''),
                contact_info=store_info.get('contact_info', ''),
                product_name=product_name,
                email_subject=email_subject,
                email_content=email_content
            )

        except Exception as e:
            logger.error(f"콜드메일 결과 자동 추가 실패: {e}")
            return False

    def _extract_store_name(self, coldmail_result: Dict[str, Any]) -> str:
        """콜드메일 결과에서 스토어명 추출"""
        # JSON 문자열에서 정보 추출 시도
        if 'body' in coldmail_result:
            content = coldmail_result['body']
            try:
                # JSON 형태로 파싱 시도
                if content.startswith('{') and 'product_name' in content:
                    data = json.loads(content.split('\n')[0])
                    return data.get('brand_name', '미확인')
            except:
                pass

        return "미확인"

    def _extract_product_name(self, coldmail_result: Dict[str, Any]) -> str:
        """콜드메일 결과에서 상품명 추출"""
        if 'body' in coldmail_result:
            content = coldmail_result['body']
            try:
                if content.startswith('{') and 'product_name' in content:
                    data = json.loads(content.split('\n')[0])
                    return data.get('product_name', '미확인')
            except:
                pass

        return "미확인"

    def _find_store_info(self, ecommerce_df: pd.DataFrame, store_name: str, product_name: str) -> Dict[str, str]:
        """이커머스 DB에서 스토어 정보 찾기"""
        store_info = {
            'store_url': '',
            'contact_info': ''
        }

        if ecommerce_df.empty:
            return store_info

        # 스토어명이나 상품명으로 매칭 시도
        matches = ecommerce_df[
            (ecommerce_df.astype(str).apply(lambda x: x.str.contains(store_name, na=False)).any(axis=1)) |
            (ecommerce_df.astype(str).apply(lambda x: x.str.contains(product_name, na=False)).any(axis=1))
        ]

        if not matches.empty:
            first_match = matches.iloc[0]
            # 가능한 컬럼명들을 시도
            url_columns = ['URL', 'url', '스토어URL', 'store_url', 'link', '링크']
            contact_columns = ['담당자', 'contact', '연락처', 'email', '이메일']

            for col in url_columns:
                if col in first_match and pd.notna(first_match[col]):
                    store_info['store_url'] = str(first_match[col])
                    break

            for col in contact_columns:
                if col in first_match and pd.notna(first_match[col]):
                    store_info['contact_info'] = str(first_match[col])
                    break

        return store_info

    def update_send_status(self, index: int, status: str, sent_datetime: Optional[datetime] = None):
        """발송 상태 업데이트"""
        try:
            df = self.load_pending_list()
            if index < len(df):
                df.loc[index, '발송상태'] = status
                if sent_datetime:
                    df.loc[index, '발송일시'] = sent_datetime.strftime('%Y-%m-%d %H:%M:%S')

                df.to_excel(self.pending_excel, index=False)
                logger.info(f"발송 상태 업데이트: 인덱스 {index}, 상태 {status}")

        except Exception as e:
            logger.error(f"발송 상태 업데이트 실패: {e}")

    def get_pending_count(self) -> int:
        """발송대기 건수 반환"""
        df = self.load_pending_list()
        return len(df[df['발송상태'] == '대기'])

    def get_today_schedule(self) -> pd.DataFrame:
        """오늘 발송 예정인 항목들 반환"""
        df = self.load_pending_list()
        today = datetime.now().strftime('%Y-%m-%d')

        return df[
            (df['발송예정일'] == today) &
            (df['발송상태'] == '대기')
        ]


if __name__ == "__main__":
    # 테스트
    manager = EmailSenderManager()

    # 샘플 데이터 추가
    manager.add_to_pending_list(
        store_name="테스트 스토어",
        store_url="https://test-store.com",
        product_name="테스트 상품",
        email_subject="상세페이지 개선 제안",
        email_content="안녕하세요. 상세페이지 분석 결과를 공유드립니다..."
    )

    print(f"발송대기 건수: {manager.get_pending_count()}")